import io

import openpyxl

from api.app.models import EvaluacionRequest
from api.services.cell_mapping import PROCESOS_ROWS
from api.services.injector import inject


def _load_result(payload):
    request = EvaluacionRequest(**payload)
    buf = inject(request)
    wb = openpyxl.load_workbook(buf, data_only=False)
    return wb["PAUTA_ERA"]


def test_inject_metadata(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["B10"].value == "CESFAM Los Alerces"
    assert ws["B11"].value == "123456"
    assert ws["B12"].value == "Santiago"
    assert ws["B13"].value == "Metropolitana"
    assert ws["B15"].value == "Juan Pérez"
    assert ws["B16"].value == "María González"
    assert ws["B17"].value == 1500
    assert ws["B18"].value == 10
    assert ws["B19"].value == "test@salud.cl"


def test_inject_tasas(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["G23"].value == 25
    assert ws["H23"].value == 5000
    assert ws["G24"].value == 15
    assert ws["H24"].value == 3000
    assert ws["G25"].value == 120
    assert ws["H25"].value == 800


def test_inject_preserves_formulas(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["I23"].value == "=G23/H23*10000"
    assert ws["I24"].value == "=G24/H24*10000"
    assert ws["I25"].value == "=G25/H25"


def test_inject_estructura_items(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["C34"].value == 1
    assert ws["C36"].value == 0
    assert ws["C38"].value == 1
    assert ws["C39"].value == 1
    assert ws["C40"].value == 0
    assert ws["C41"].value == 1
    assert ws["C42"].value == 1
    assert ws["C43"].value == 0
    assert ws["C44"].value == 1
    assert ws["C45"].value == 1
    assert ws["C46"].value == 1
    assert ws["C47"].value == 0


def test_inject_procesos_items(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["C52"].value == 1
    assert ws["C53"].value == 1
    assert ws["C54"].value == 0
    assert ws["C55"].value == 1
    assert ws["C56"].value == 0
    assert ws["C57"].value == 1
    assert ws["C58"].value == 1
    assert ws["C59"].value == 1
    assert ws["C61"].value == 0
    assert ws["C62"].value == 1
    assert ws["C63"].value == 1
    assert ws["C64"].value == 0
    assert ws["C65"].value == 1
    assert ws["C66"].value == 1
    assert ws["C68"].value == 0
    assert ws["C69"].value == 1
    assert ws["C70"].value == 1
    assert ws["C71"].value == 1
    assert ws["C73"].value == 0
    assert ws["C74"].value == 1
    assert ws["C76"].value == 1


def test_inject_compromisos(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["A97"].value == "Implementar protocolo de rescate de inasistentes"


def test_inject_estructura_observations(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["J34"].value == "Box disponible"
    assert ws["J40"].value == "Sin espirómetro"
    assert ws["J47"].value == "Sin sistema de citas"


def test_inject_procesos_observations(sample_payload):
    ws = _load_result(sample_payload)
    for item in sample_payload["evaluacion"]["procesos"]:
        if item["observacion"]:
            row = PROCESOS_ROWS[item["item"]]
            assert ws[f"J{row}"].value == item["observacion"]


def test_preserves_summary_formulas(sample_payload):
    ws = _load_result(sample_payload)
    assert ws["C81"].value == "=COUNT(C34,C36,C38:C47)"
    assert ws["C82"].value == "=COUNT(C76,C74,C73,C71,C70,C69,C68,C66,C65,C64,C63,C62,C61,C59,C58,C57,C56,C55,C54,C53,C52)"
    assert ws["D81"].value == "=SUM(C34:C47)"
    assert ws["D82"].value == "=SUM(C52:C76)"
    assert ws["F81"].value == "=D81/C81"
    assert ws["F82"].value == "=D82/C82"
    assert ws["I81"].value == '=IF(F81<0.6,"POR LOGRAR",IF(AND(F81<0.9,F81>=0.6),"PARCIALMENTE LOGRADO","LOGRADO"))'
    assert ws["I82"].value == '=IF(F82<0.6,"POR LOGRAR",IF(AND(F82<0.9,F82>=0.6),"PARCIALMENTE LOGRADO","LOGRADO"))'
