import io
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from api.app.models import EvaluacionItem, EvaluacionRequest
from api.services.cell_mapping import (
    COMPROMISOS_CELL,
    ESTRUCTURA_ROWS,
    FORMULA_CELLS,
    METADATA_CELLS,
    PROCESOS_ROWS,
    TASAS_CELLS,
)

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "plantilla_base.xlsx"


def _safe_write(ws: Worksheet, cell_ref: str, value: object) -> None:
    cell = ws[cell_ref]
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for mr in list(ws.merged_cells.ranges):
            if cell.coordinate in mr:
                ws.unmerge_cells(str(mr))
                break
    ws[cell_ref] = value


def _inject_items(ws: Worksheet, items: list[EvaluacionItem], row_map: dict[int, int]) -> None:
    for item in items:
        row = row_map.get(item.item)
        if row:
            _safe_write(ws, f"C{row}", item.puntaje)
            if item.observacion:
                _safe_write(ws, f"J{row}", item.observacion)


def inject(payload: EvaluacionRequest) -> io.BytesIO:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["PAUTA_ERA"]

    formulas_before = {cell: ws[cell].value for cell in FORMULA_CELLS}

    for field, cell in METADATA_CELLS.items():
        _safe_write(ws, cell, getattr(payload.metadata, field))

    for tasa_name, cells in TASAS_CELLS.items():
        tasa = getattr(payload.tasas_resultado, tasa_name)
        _safe_write(ws, cells["numerador"], tasa.numerador)
        _safe_write(ws, cells["denominador"], tasa.denominador)

    _inject_items(ws, payload.evaluacion.estructura, ESTRUCTURA_ROWS)
    _inject_items(ws, payload.evaluacion.procesos, PROCESOS_ROWS)

    _safe_write(ws, COMPROMISOS_CELL, payload.cierre.compromisos)

    for cell, original in formulas_before.items():
        if ws[cell].value != original:
            ws[cell] = original

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
